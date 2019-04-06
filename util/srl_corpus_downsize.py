# -*- coding:utf-8 -*-
from pyltp import SementicRoleLabeller
from pyltp import Segmentor
from pyltp import Postagger
from pyltp import Parser
from matplotlib import pyplot as plt
import openpyxl
import codecs
import os
import math
import random

'''存储单个文件的结果信息'''
class DataInfo():
    def __init__(self,
                 predicate_arcmeaning_ratio,
                 predicate_postags_ratio,
                 predicate_words_ratio,
                 arg_argmeaning_ratio,
                 arg_arcs_ratio,
                 arg_postags_ratio,
                 arg_argmeaning_arcsratiodict,
                 arg_argmeaning_postagsratiodict):
        self.predicate_arcmeaning_ratio = predicate_arcmeaning_ratio
        self.predicate_postags_ratio = predicate_postags_ratio
        self.predicate_words_ratio = predicate_words_ratio
        self.arg_argmeaning_ratio = arg_argmeaning_ratio
        self.arg_arcs_ratio = arg_arcs_ratio
        self.arg_postags_ratio = arg_postags_ratio
        self.arg_argmeaning_arcs_ratio = arg_argmeaning_arcsratiodict
        self.arg_argmeaning_postags_ratio = arg_argmeaning_postagsratiodict

'''存储文件夹下所有文件的结果信息'''
class DataDict():
    dict_filename_predicate_arcmeaning_ratio = {}
    dict_filename_predicate_postags_ratio = {}
    dict_filename_predicate_words_ratio = {}
    dict_filename_arg_argmeaning_ratio = {}
    dict_filename_arg_arcs_ratio = {}
    dict_filename_arg_postags_ratio = {}
    dict_filename_arg_argmeaning_arcsratiodict = {}
    dict_filename_arg_argmeaning_postagsratiodict = {}
    def __init__(self,
                 filename,
                 predicate_arcmeaning_ratio,
                 predicate_postags_ratio,
                 predicate_words_ratio,
                 arg_argmeaning_ratio,
                 arg_arcs_ratio,
                 arg_postags_ratio,
                 arg_argmeaning_arcsratiodict,
                 arg_argmeaning_postagsratiodict):
        if filename not in self.dict_filename_predicate_arcmeaning_ratio:
            self.dict_filename_predicate_arcmeaning_ratio.update({filename : predicate_arcmeaning_ratio})

        if filename not in self.dict_filename_predicate_postags_ratio:
            self.dict_filename_predicate_postags_ratio.update({filename : predicate_postags_ratio})

        if filename not in self.dict_filename_predicate_words_ratio:
            self.dict_filename_predicate_words_ratio.update({filename : predicate_words_ratio})

        if filename not in self.dict_filename_arg_argmeaning_ratio:
            self.dict_filename_arg_argmeaning_ratio.update({filename : arg_argmeaning_ratio})

        if filename not in self.dict_filename_arg_arcs_ratio:
            self.dict_filename_arg_arcs_ratio.update({filename : arg_arcs_ratio})

        if filename not in self.dict_filename_arg_postags_ratio:
            self.dict_filename_arg_postags_ratio.update({filename : arg_postags_ratio})

        if filename not in self.dict_filename_arg_argmeaning_arcsratiodict:
            self.dict_filename_arg_argmeaning_arcsratiodict.update({filename : arg_argmeaning_arcsratiodict})

        if filename not in self.dict_filename_arg_argmeaning_postagsratiodict:
            self.dict_filename_arg_argmeaning_postagsratiodict.update({filename : arg_argmeaning_postagsratiodict})

class SRL():
    # ltp的模型文件
    LTP_DATA_DIR = "E:\ltp_model"
    dict_SRL = {
        #主要角色(参考The Proposition Bank: An Annotated Corpus of Semantic Roles)
        "A0" : "施事",#一般为动作方
        "A1" : "受事",#一般为动作的影响
        "A2" : "范围",#具体含义根据谓词的含义有变化
        "A3" : "动作开始",#具体含义根据谓词的含义有变化
        "A4" : "动作结束",#具体含义根据谓词的含义有变化
        "A5" : "其他动作相关",#具体含义根据谓词的含义有变化

        #附加角色
        "ADV" : "附加|默认标记",
        "BNE" : "受益人",
        "BNF" : "受益人",
        "CND" : "条件",
        "DIR" : "方向",
        "DGR" : "程度",
        "EXT" : "范围",
        "FRQ" : "频率",
        "LOC" : "地点",
        "MNR" : "方式",
        "PRP" : "目的|原因",
        "TMP" : "时间",
        "TPC" : "主题",
        "DIS": "标记语",
        "CRD" : "并列参数",
        "PRD" : "谓语动词",
        "PSR" : "持有者",
        "PSE" : "被持有"
    }

    dict_SRL_reverse = {
        # 主要角色(参考The Proposition Bank: An Annotated Corpus of Semantic Roles)
        "施事": "A0",  # 一般为动作方
        "受事": "A1",  # 一般为动作的影响
        "范围相关": "A2",  # 具体含义根据谓词的含义有变化
        "动作开始": "A3",  # 具体含义根据谓词的含义有变化
        "动作结束": "A4",  # 具体含义根据谓词的含义有变化
        "其他动作相关": "A5",  # 具体含义根据谓词的含义有变化

        # 附加角色
        "附加|默认标记": "ADV",
        "受益人": "BNF",
        "条件": "CND",
        "方向": "DIR",
        "程度": "DGR",
        "范围": "EXT",
        "频率": "FRQ",
        "地点": "LOC",
        "方式": "MNR",
        "目的|原因": "PRP",
        "时间": "TMP",
        "主题": "TPC",
        "标记语": "DIS",
        "并列参数": "CRD",
        "谓语动词": "PRD",
        "持有者": "PSR",
        "被持有": "PSE"
    }

    dictSyntax = {
        'SBV': '主谓关系',
        'VOB': '动宾关系',
        'IOB': '间宾关系',
        'FOB': '前置宾语',
        'DBL': '兼语',
        'ATT': '定中关系',
        'ADV': '状中结构',
        'CMP': '动补结构',
        'COO': '并列关系',
        'POB': '介宾关系',
        'LAD': '左附加关系',
        'RAD': '右附加关系',
        'IS': '独立结构',
        'HED': '核心关系',
        'WP': '标点关系',
    }

    def readfilelist(self, path):
        list_corpusfilename = os.listdir(path)
        dict_path_corpus = {}
        for i in range(len(list_corpusfilename)):
            filepath = os.path.join(path, list_corpusfilename[i])
            key = list_corpusfilename[i]
            value = self.readfile(filepath)
            dict_path_corpus.update({key:value})
        return dict_path_corpus

    def readfile(self, path):
        with open(path, 'r', encoding='utf-8') as file_corpus:
            list_corpus = file_corpus.readlines()
        for i in range(len(list_corpus)):
            line = list_corpus[i].strip('\n')
            list_corpus[i] = line
        return list_corpus

    '''分词模块'''
    def word_segmentation(self, list_corpus):
        list_word_segmentation = []
        cws_model_path = os.path.join(self.LTP_DATA_DIR, 'cws.model')
        segmentor = Segmentor()
        segmentor.load(cws_model_path)
        for sentence in list_corpus:
            words = segmentor.segment(sentence)
            list_words = list(words)
            list_word_segmentation.append(list_words)
        segmentor.release()
        return list_word_segmentation
    '''词性标注模块'''
    def postagger(self, list_word_segmentation):
        list_postags = []
        pos_model_path = os.path.join(self.LTP_DATA_DIR, 'pos.model')
        postagger = Postagger()  # 初始化实例
        postagger.load(pos_model_path)  # 加载模型
        for list_word in list_word_segmentation:
            postags = postagger.postag(list_word)
            postags = list(postags)
            list_postags.append(postags)
        postagger.release()
        return list_postags

    '''句法分析模块'''
    def syntax_parser(self, list_word_segmentation, list_postags):
        list_arcs = []
        list_arcs_content = []
        par_model_path = os.path.join(self.LTP_DATA_DIR, 'parser.model')
        parser = Parser()  # 初始化实例
        parser.load(par_model_path)  # 加载模型
        for i in range(len(list_word_segmentation)):
            arcs = parser.parse(list_word_segmentation[i], list_postags[i])
            list_arc_temp = []
            for arc in arcs:
                list_arc_temp.append(arc.relation)
            list_arcs_content.append(list_arc_temp)
            list_arcs.append(arcs)
            # print("|".join("%d:%s" % (arc.head, arc.relation) for arc in arcs))
        parser.release()
        return list_arcs, list_arcs_content

    '''语义角色标注模块'''
    def sementic_role_label(self, list_word_segmentation, list_postags, list_arcs):
        list_roles = []
        srl_model_path = os.path.join(self.LTP_DATA_DIR, 'pisrl_win.model')
        labeller = SementicRoleLabeller()
        labeller.load(srl_model_path)
        for i in range(len(list_word_segmentation)):
            roles = labeller.label(list_word_segmentation[i], list_postags[i], list_arcs[i])
            list_roles.append(roles)
            # for role in roles:
                # print(str(role.index), "".join(
                #     ["%s:(%d,%d)" % (arg.name, arg.range.start, arg.range.end) for arg in role.arguments]))
        return list_roles

    '''语义角色标注结果格式化'''
    def formatSRL(self, list_roles):
        # dict_sentenceindex_roleindex_argdict
        # {s1:{
        #      index1: {A0:[[start, end]], ADV:[[start, end], [start, end]]},
        #      index2: {A0:[[start, end]], ADV:[[start, end], [start, end]]},
        #     ...
        #    }
        # s2:{
        #      index1: {A0:[[start, end]], ADV:[[start, end], [start, end]]},
        #      index2: {A0:[[start, end]], ADV:[[start, end], [start, end]]},
        #      ...
        #     }
        # }
        dict_sentenceindex_roleindex_argdict = {}
        for i in range(len(list_roles)):
            sentence_roles = list_roles[i]
            dict_roleindex_argdict = {}
            for role in sentence_roles:
                dict_rolename_scope = {}
                roleindex = role.index
                for arg in role.arguments:
                    name = arg.name
                    if name in dict_rolename_scope:
                        list_temp = dict_rolename_scope[name]
                        list_temp.append([int(arg.range.start), int(arg.range.end) + 1])
                        dict_rolename_scope.update({arg.name : list_temp})
                    else:
                        dict_rolename_scope.update({arg.name : [[int(arg.range.start), int(arg.range.end) + 1]]})
                dict_roleindex_argdict.update({roleindex : dict_rolename_scope})
            dict_sentenceindex_roleindex_argdict.update({i : dict_roleindex_argdict})
        return dict_sentenceindex_roleindex_argdict
        pass

    '''根据语义角色标注的位置信息，解析为具体词语'''
    def parseSRL(self, dict_sentenceindex_roleindex_argdict, list_word_segmentation, list_postags, list_arcs):
        # 中间roleindex只是下标，如果想要找到词，则要根据分词的结果list去找，几个不同的dict都是其他内容有变化
        list_core_words = []
        list_core_words_postags = []
        list_core_words_arcs = []
        list_core_words_arcs_meaning = []

        dict_sentenceindex_roleindex_argmeaningwordsdict = {}
        dict_sentenceindex_roleindex_argmeaningpostagsdict = {}
        dict_sentenceindex_roleindex_argmeaningarcsdict = {}

        for sentenceindex, roles in dict_sentenceindex_roleindex_argdict.items():
            sentence_word_segmentation = list_word_segmentation[sentenceindex]
            sentence_postags = list_postags[sentenceindex]
            sentence_arcs = list_arcs[sentenceindex]

            dict_roleindex_argmeaningwordsdict = {}
            dict_roleindex_argmeaningpostagsdict = {}
            dict_roleindex_argmeaningarcsdict = {}

            temp_list_core_words = []
            temp_list_core_words_postags = []
            temp_list_core_words_arcs = []
            temp_list_core_words_arcs_meaning = []
            for roleindex, role in roles.items():
                core_word = sentence_word_segmentation[roleindex]
                core_word_postag = sentence_postags[roleindex]
                core_word_arc = sentence_arcs[roleindex]
                core_word_arc_meaning = self.dictSyntax[core_word_arc]

                temp_list_core_words.append(core_word)
                temp_list_core_words_postags.append(core_word_postag)
                temp_list_core_words_arcs.append(core_word_arc)
                temp_list_core_words_arcs_meaning.append(core_word_arc_meaning)

                dict_argmeaning_wordsscope = {}
                dict_argmeaning_postagsscope = {}
                dict_argmeaning_arcsscope = {}
                for arg_name, list_args in role.items():
                    if arg_name in self.dict_SRL:
                        arg_meaning = self.dict_SRL[arg_name]
                    else:
                        arg_meaning = arg_name
                    for scope in list_args:
                        lower = scope[0]
                        upper = scope[1]
                        scope_words = sentence_word_segmentation[lower:upper]
                        scope_words_postags = sentence_postags[lower:upper]
                        scope_words_arcs = sentence_arcs[lower:upper]

                        if arg_meaning in dict_argmeaning_postagsscope:
                            list_temp = dict_argmeaning_postagsscope[arg_meaning]
                            list_temp.append(scope_words_postags)
                            dict_argmeaning_postagsscope.update({arg_meaning : list_temp})
                        else:
                            dict_argmeaning_postagsscope.update({arg_meaning : [scope_words_postags]})

                        if arg_meaning in dict_argmeaning_arcsscope:
                            list_temp = dict_argmeaning_arcsscope[arg_meaning]
                            list_temp.append(scope_words_arcs)
                            dict_argmeaning_arcsscope.update({arg_meaning : list_temp})
                        else:
                            dict_argmeaning_arcsscope.update({arg_meaning : [scope_words_arcs]})
                            pass

                        if arg_meaning in dict_argmeaning_wordsscope:
                            list_temp = dict_argmeaning_wordsscope[arg_meaning]
                            list_temp.append(scope_words)
                            dict_argmeaning_wordsscope.update({arg_meaning : list_temp})
                        else:
                            dict_argmeaning_wordsscope.update({arg_meaning : [scope_words]})

                dict_roleindex_argmeaningwordsdict.update({roleindex : dict_argmeaning_wordsscope})
                dict_roleindex_argmeaningpostagsdict.update({roleindex : dict_argmeaning_postagsscope})
                dict_roleindex_argmeaningarcsdict.update({roleindex : dict_argmeaning_arcsscope})

            list_core_words.append(temp_list_core_words)
            list_core_words_postags.append(temp_list_core_words_postags)
            list_core_words_arcs.append(temp_list_core_words_arcs)
            list_core_words_arcs_meaning.append(temp_list_core_words_arcs_meaning)

            dict_sentenceindex_roleindex_argmeaningwordsdict.update({sentenceindex : dict_roleindex_argmeaningwordsdict})
            dict_sentenceindex_roleindex_argmeaningpostagsdict.update({sentenceindex : dict_roleindex_argmeaningpostagsdict})
            dict_sentenceindex_roleindex_argmeaningarcsdict.update({sentenceindex : dict_roleindex_argmeaningarcsdict})
        return dict_sentenceindex_roleindex_argmeaningwordsdict, \
               dict_sentenceindex_roleindex_argmeaningpostagsdict, \
               dict_sentenceindex_roleindex_argmeaningarcsdict, \
               list_core_words, list_core_words_postags, \
               list_core_words_arcs, \
               list_core_words_arcs_meaning

    '''统计模块'''
    def stasticalSRL(self, dict_sentenceindex_roleindex_argmeaningwordsdict, dict_sentenceindex_roleindex_argmeaningpostagsdict, dict_sentenceindex_roleindex_argmeaningarcsdict, list_core_words, list_core_words_postags, list_core_words_arcs, list_core_words_arcs_meaning):
        # arc代表语法树角色标签，arcmeaning代表语法树角色标签含义，postag代表词性标签
        # 主结构谓词(core word)统计
        dict_arc_words = {}
        dict_word_arcs = {}
        dict_postag_words = {}
        dict_word_postags = {}

        dict_arc_wordscount = {}
        dict_postag_wordscount = {}
        dict_word_count = {}
        for i in range(len(list_core_words)):
            sentence_core_words = list_core_words[i]
            sentence_core_words_postags = list_core_words_postags[i]
            sentence_core_words_arcs = list_core_words_arcs[i]
            sentence_core_words_arcsmeaning = list_core_words_arcs_meaning[i]

            for j in range(len(sentence_core_words)):
                word = sentence_core_words[j]
                word_postag = sentence_core_words_postags[j]
                word_arc = sentence_core_words_arcs[j]
                word_arcmeaning = sentence_core_words_arcsmeaning[j]
                if word_arc in dict_arc_words:
                    list_temp = dict_arc_words[word_arc]
                    list_temp.append(word)
                    dict_arc_words.update({word_arc : list_temp})
                else:
                    dict_arc_words.update({word_arc : [word]})

                if word_arcmeaning in dict_arc_wordscount:
                    amount = dict_arc_wordscount[word_arcmeaning]
                    amount = amount + 1
                    dict_arc_wordscount.update({word_arcmeaning : amount})
                else:
                    dict_arc_wordscount.update({word_arcmeaning : 1})

                if word_postag in dict_postag_words:
                    list_temp = dict_postag_words[word_postag]
                    list_temp.append(word)
                    dict_postag_words.update({word_postag : list_temp})
                else:
                    dict_postag_words.update({word_postag : [word]})

                if word_postag in dict_postag_wordscount:
                    amount = dict_postag_wordscount[word_postag]
                    amount = amount + 1
                    dict_postag_wordscount.update({word_postag : amount})
                else:
                    dict_postag_wordscount.update({word_postag : 1})

                if word in dict_word_arcs:
                    list_temp = dict_word_arcs[word]
                    list_temp.append(word_arc)
                    dict_word_arcs.update({word : list_temp})

                    list_temp = dict_word_postags[word]
                    list_temp.append(word_postag)
                    dict_word_postags.update({word : list_temp})
                else:
                    dict_word_arcs.update({word : [word_arc]})
                    dict_word_postags.update({word : [word_postag]})

                if word in dict_word_count:
                    amount = dict_word_count[word]
                    amount = amount + 1
                    dict_word_count.update({word: amount})

                    # amount = dict_word_postagscount[word]
                    # amount = amount + 1
                    # dict_word_postagscount.update({word : amount})
                else:
                    dict_word_count.update({word: 1})
                    # dict_word_postagscount.update({word : 1})

        # 语义角色标签下的各种结构下的语法结构比例和词性比例
        dict_argmeaning_allwords = {}
        dict_argmeaning_allarcs = {}
        dict_argmeaning_allpostags = {}
        dict_argmeaning_wordscount = {}
        # dict_argmeaning_allpostagscount = {}

        dict_arc_amount = {} # 对于具有语义角色成分范围的全集语法结构比例
        dict_postag_amount = {} # 对于具有语义角色成分范围的全集词性比例

        dict_argmeaning_arcscountdict = {} # 语义角色标签下的各类语法树角色计数
        dict_argmeaning_postagscountdict = {} # 语义角色标签下的各类词性计数
        for i in range(len(dict_sentenceindex_roleindex_argmeaningwordsdict)):
            sentence = list_word_segmentation[i]
            roleindex_argmeaningwordsdict = dict_sentenceindex_roleindex_argmeaningwordsdict[i]
            roleindex_argmeaningarcsdict = dict_sentenceindex_roleindex_argmeaningarcsdict[i]
            roleindex_argmeaningpostagsdict = dict_sentenceindex_roleindex_argmeaningpostagsdict[i]

            for roleindex in roleindex_argmeaningwordsdict:
                dict_argmeaning_words = roleindex_argmeaningwordsdict[roleindex]
                dict_argmeaning_arcs = roleindex_argmeaningarcsdict[roleindex]
                dict_argmeaning_postags = roleindex_argmeaningpostagsdict[roleindex]
                for argmeaning in dict_argmeaning_words:
                    list_words = dict_argmeaning_words[argmeaning]
                    list_arcs = dict_argmeaning_arcs[argmeaning]
                    list_postags = dict_argmeaning_postags[argmeaning]
                    # 填充dict_arc_amount和dict_postag_amount
                    for arcs in list_arcs:
                        for arc in arcs:
                            if arc in dict_arc_amount:
                                amount = dict_arc_amount[arc]
                                amount = amount + 1
                                dict_arc_amount.update({arc: amount})
                            else:
                                amount = 1
                                dict_arc_amount.update({arc: amount})

                    for postags in list_postags:
                        for postag in postags:
                            if postag in dict_postag_amount:
                                amount = dict_postag_amount[postag]
                                amount = amount + 1
                                dict_postag_amount.update({postag: amount})
                            else:
                                amount = 1
                                dict_postag_amount.update({postag: amount})


                    # 填充dict_argmeaning_allarcs和dict_argmeaning_allpostags
                    list_arcsall = []
                    list_postagsall = []
                    list_wordsall = []
                    for arcs in list_arcs:
                        for arc in arcs:
                            list_arcsall.append(arc)
                    for postags in list_postags:
                        for postag in postags:
                            list_postagsall.append(postag)
                    for words in list_words:
                        for word in words:
                            list_wordsall.append(word)
                    if argmeaning in dict_argmeaning_allarcs:
                        list_temp = dict_argmeaning_allarcs[argmeaning]
                        for arcs in list_arcs:
                            list_temp = list_temp + arcs
                        # list_temp = list_temp + list_arcs
                        dict_argmeaning_allarcs.update({argmeaning: list_temp})

                        length = len(list_temp)
                        dict_argmeaning_wordscount.update({argmeaning : length})

                        list_temp = dict_argmeaning_allpostags[argmeaning]
                        for postags in list_postags:
                            list_temp = list_temp + postags
                        dict_argmeaning_allpostags.update({argmeaning : list_temp})

                        # length = len(list_temp)
                        # dict_argmeaning_allpostagscount.update({argmeaning : length})
                    else:
                        dict_argmeaning_allarcs.update({argmeaning: list_arcsall})
                        dict_argmeaning_allpostags.update({argmeaning : list_postagsall})
                        dict_argmeaning_wordscount.update({argmeaning : len(list_arcsall)})
                        # dict_argmeaning_allpostagscount.update({argmeaning : len(list_postags)})
                    # 填充dict_argmeaning_arcscountdict和dict_argmeaning_postagscountdict
                    # {
                    #   argmeaning : {arc1 : 3, arc2 : 2}
                    # }
                    if argmeaning in dict_argmeaning_arcscountdict:
                        dict_arcs_amount = dict_argmeaning_arcscountdict[argmeaning]
                        for j in range(len(list_wordsall)):
                            word = list_wordsall[j]
                            arc = list_arcsall[j]
                            # if len(list_arcsall)!=len(list_wordsall):
                            #     print(list_wordsall)
                            #     print('\n')
                            #     print(list_arcsall)
                            #     print('\n')
                            #     print(i)
                            #     print('\n')
                            #     print(roleindex)
                            #     print('\n')
                            #     print(argmeaning)
                            #     print('\n')
                            if arc in dict_arcs_amount:
                                amount = dict_arcs_amount[arc]
                                amount = amount + 1
                                dict_arcs_amount.update({arc : amount})
                            else:
                                amount = 1
                                dict_arcs_amount.update({arc : amount})
                        dict_argmeaning_arcscountdict.update({argmeaning : dict_arcs_amount})
                    else:
                        dict_arcs_amount = {}
                        for j in range(len(list_wordsall)):
                            arc = list_arcsall[j]
                            dict_arcs_amount.update({arc : 1})
                        dict_argmeaning_arcscountdict.update({argmeaning : dict_arcs_amount})

                    if argmeaning in dict_argmeaning_postagscountdict:
                        dict_postags_amount = dict_argmeaning_postagscountdict[argmeaning]
                        for j in range(len(list_wordsall)):
                            postag = list_postagsall[j]
                            if postag in dict_postags_amount:
                                amount = dict_postags_amount[postag]
                                amount = amount + 1
                                dict_postags_amount.update({postag : amount})
                            else:
                                dict_postags_amount.update({postag : 1})
                        dict_argmeaning_postagscountdict.update({argmeaning : dict_postags_amount})
                    else:
                        dict_postags_amount = {}
                        for j in range(len(list_wordsall)):
                            postag = list_postagsall[j]
                            dict_postags_amount.update({postag : 1})
                        dict_argmeaning_postagscountdict.update({argmeaning : dict_postags_amount})
        return dict_arc_wordscount, dict_word_count, dict_postag_wordscount, dict_argmeaning_wordscount, dict_arc_amount, dict_postag_amount, dict_argmeaning_arcscountdict, dict_argmeaning_postagscountdict
        pass

    # 打印三重字典
    def writeSRL(self, dict_sentenceindex_roleindex_argmeaningwordsdict, list_word_segmentation, path):
        file_write = codecs.open(path, 'w', 'utf-8')
        for i in range(len(dict_sentenceindex_roleindex_argmeaningwordsdict)):
            sentence_words = list_word_segmentation[i]
            dict_roleindex_argmeaningwordsdict = dict_sentenceindex_roleindex_argmeaningwordsdict[i]
            for roleindex, argmeaningwordsdict in dict_roleindex_argmeaningwordsdict:
                words_predicate = sentence_words[roleindex]
                file_write.write(words_predicate + " : " + str(argmeaningwordsdict) + " ||| ")
            file_write.write(os.linesep)
        file_write.flush()
        file_write.close()

    # 计算各种比例
    def ratio_predicate(self, dict_arc_wordscount, dict_postag_wordscount, dict_word_count):
        total = 0
        dict_arcmeaning_ratio = {}
        for arcmeaning, amount in dict_arc_wordscount.items():
            total = total + amount
        for arcmeaning, amount in dict_arc_wordscount.items():
            ratio = amount/total
            dict_arcmeaning_ratio.update({arcmeaning : [amount, ratio]})
        dict_arcmeaning_ratio = sorted(dict_arcmeaning_ratio.items(), key=lambda item:item[1][0], reverse=True)

        dict_postags_ratio = {}
        for postag, amount in dict_postag_wordscount.items():
            ratio = amount/total
            dict_postags_ratio.update({postag : [amount, ratio]})
        dict_postags_ratio = sorted(dict_postags_ratio.items(), key=lambda item:item[1][0], reverse=True)

        dict_word_ratio = {}
        for word, amount in dict_word_count.items():
            ratio = amount/total
            dict_word_ratio.update({word : [amount , ratio]})
        dict_word_ratio = sorted(dict_word_ratio.items(), key=lambda item:item[1][0], reverse=True)
        return dict_arcmeaning_ratio, dict_postags_ratio, dict_word_ratio
        pass

    def ratio_argument(self, dict_argmeaning_wordscount, dict_arc_amount, dict_postag_amount, dict_argmeaning_arcscountdict, dict_argmeaning_postagscountdict):
        total = 0
        for argmeaning, wordscount in dict_argmeaning_wordscount.items():
            total = total + wordscount

        dict_argmeaning_ratio = {}
        for argmeaning, wordscount in dict_argmeaning_wordscount.items():
            ratio = wordscount/total
            dict_argmeaning_ratio.update({argmeaning : [wordscount, ratio]})
        dict_argmeaning_ratio = sorted(dict_argmeaning_ratio.items(), key=lambda item: item[1][0], reverse=True)

        dict_arc_ratio = {}
        for arc, amount in dict_arc_amount.items():
            ratio = amount/total
            dict_arc_ratio.update({arc : [amount, ratio]})
        dict_arc_ratio = sorted(dict_arc_ratio.items(), key=lambda item: item[1][0], reverse=True)

        dict_postag_ratio = {}
        for postag, amount in dict_postag_amount.items():
            ratio = amount/total
            dict_postag_ratio.update({postag : [amount, ratio]})
        dict_postag_ratio = sorted(dict_postag_ratio.items(), key=lambda item: item[1][0], reverse=True)

        dict_argmeaning_arcsratiodict = {}
        for argmeaning, arcscountdict in dict_argmeaning_arcscountdict.items():
            total_inner = 0
            dict_arcs_ratio = {}
            for arc, amount in arcscountdict.items():
                total_inner = total_inner + amount
            for arc, amount in arcscountdict.items():
                ratio = amount/total_inner
                dict_arcs_ratio.update({arc : [amount, ratio]})
            dict_arcs_ratio = sorted(dict_arcs_ratio.items(), key=lambda item: item[1][0], reverse=True)
            dict_argmeaning_arcsratiodict.update({argmeaning : dict_arcs_ratio})
        dict_argmeaning_arcsratiodict = sorted(dict_argmeaning_arcsratiodict.items(), key=lambda item: item[0], reverse=True)

        dict_argmeaning_postagsratiodict = {}
        for argmeaning, postagscountdict in dict_argmeaning_postagscountdict.items():
            total_inner = 0
            dict_postags_ratio = {}
            for postag, amount in postagscountdict.items():
                total_inner = total_inner + amount
            for postag, amount in postagscountdict.items():
                ratio = amount/total_inner
                dict_postags_ratio.update({postag : [amount, ratio]})
            dict_postags_ratio = sorted(dict_postags_ratio.items(), key=lambda item: item[1][0], reverse=True)
            dict_argmeaning_postagsratiodict.update({argmeaning : dict_postags_ratio})
        dict_argmeaning_postagsratiodict = sorted(dict_argmeaning_postagsratiodict.items(), key=lambda item: item[0], reverse=True)
        return dict_argmeaning_ratio, dict_arc_ratio, dict_postag_ratio, dict_argmeaning_arcsratiodict, dict_argmeaning_postagsratiodict


    def writeXLS(self, datainfo, list_origion_text,
                 dict_sentenceindex_roleindex_argmeaningwordsdict,
                 dict_sentenceindex_roleindex_argmeaningpostagsdict,
                 dict_sentenceindex_roleindex_argmeaningarcsdict,
                 list_word_segmentation,
                 list_postags,
                 list_arcs_content,
                 list_contribute,
                 filename):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "谓词语法树角色比例"
        predicate_arcmeaning_ratio = datainfo.predicate_arcmeaning_ratio
        for i in range(len(predicate_arcmeaning_ratio)):
            sheet.cell(row=i+1, column=1, value=predicate_arcmeaning_ratio[i][0])
            sheet.cell(row=i+1, column=2, value=predicate_arcmeaning_ratio[i][1][0])
            sheet.cell(row=i+1, column=3, value=predicate_arcmeaning_ratio[i][1][1])

        sheet1 = wb.create_sheet()
        sheet1.title = "谓词词性比例"
        predicate_postags_ratio = datainfo.predicate_postags_ratio
        for i in range(len(predicate_postags_ratio)):
            sheet1.cell(row=i + 1, column=1, value=predicate_postags_ratio[i][0])
            sheet1.cell(row=i + 1, column=2, value=predicate_postags_ratio[i][1][0])
            sheet1.cell(row=i + 1, column=3, value=predicate_postags_ratio[i][1][1])

        sheet2 = wb.create_sheet()
        sheet2.title = "谓词数目比例"
        predicate_words_ratio = datainfo.predicate_words_ratio
        for i in range(len(predicate_words_ratio)):
            sheet2.cell(row=i + 1, column=1, value=predicate_words_ratio[i][0])
            sheet2.cell(row=i + 1, column=2, value=predicate_words_ratio[i][1][0])
            sheet2.cell(row=i + 1, column=3, value=predicate_words_ratio[i][1][1])

        sheet3 = wb.create_sheet(title="语义角色比例")
        arg_argmeaning_ratio = datainfo.arg_argmeaning_ratio
        for i in range(len(arg_argmeaning_ratio)):
            sheet3.cell(row=i + 1, column=1, value=arg_argmeaning_ratio[i][0])
            sheet3.cell(row=i + 1, column=2, value=arg_argmeaning_ratio[i][1][0])
            sheet3.cell(row=i + 1, column=3, value=arg_argmeaning_ratio[i][1][1])

        sheet4 = wb.create_sheet(title="语义角色范围的语法树角色比例")
        arg_arcs_ratio = datainfo.arg_arcs_ratio
        for i in range(len(arg_arcs_ratio)):
            sheet4.cell(row=i + 1, column=1, value=arg_arcs_ratio[i][0])
            sheet4.cell(row=i + 1, column=2, value=arg_arcs_ratio[i][1][0])
            sheet4.cell(row=i + 1, column=3, value=arg_arcs_ratio[i][1][1])

        sheet5 = wb.create_sheet(title="语义角色范围的词性比例")
        arg_postags_ratio = datainfo.arg_postags_ratio
        for i in range(len(arg_postags_ratio)):
            sheet5.cell(row=i + 1, column=1, value=arg_postags_ratio[i][0])
            sheet5.cell(row=i + 1, column=2, value=arg_postags_ratio[i][1][0])
            sheet5.cell(row=i + 1, column=3, value=arg_postags_ratio[i][1][1])

        sheet6 = wb.create_sheet(title="语义类内的语义角色比例")
        arg_argmeaning_arcs_ratio = datainfo.arg_argmeaning_arcs_ratio
        row_count = 0
        for index_argmeaning in range(len(arg_argmeaning_arcs_ratio)):
            argmeaning_arcs_ratio = arg_argmeaning_arcs_ratio[index_argmeaning]
            argmeaning = argmeaning_arcs_ratio[0]
            list_arcs_ratio = argmeaning_arcs_ratio[1]
            for i in range(len(list_arcs_ratio)):
                arc = list_arcs_ratio[i][0]
                amount = list_arcs_ratio[i][1][0]
                ratio = list_arcs_ratio[i][1][1]
                row_count = row_count + 1
                sheet6.cell(row=row_count, column=1, value=argmeaning)
                sheet6.cell(row=row_count, column=2, value=arc)
                sheet6.cell(row=row_count, column=3, value=amount)
                sheet6.cell(row=row_count, column=4, value=ratio)

        sheet7 = wb.create_sheet(title="语义类内词性比例")
        arg_argmeaning_postags_ratio = datainfo.arg_argmeaning_postags_ratio
        row_count = 0
        for index_argmeaning in range(len(arg_argmeaning_postags_ratio)):
            argmeaning_postags_ratio = arg_argmeaning_postags_ratio[index_argmeaning]
            argmeaning = argmeaning_postags_ratio[0]
            list_postags_ratio = argmeaning_postags_ratio[1]
            for i in range(len(list_postags_ratio)):
                postag = list_postags_ratio[i][0]
                amount = list_postags_ratio[i][1][0]
                ratio = list_postags_ratio[i][1][1]
                row_count = row_count + 1
                sheet7.cell(row=row_count, column=1, value=argmeaning)
                sheet7.cell(row=row_count, column=2, value=postag)
                sheet7.cell(row=row_count, column=3, value=amount)
                sheet7.cell(row=row_count, column=4, value=ratio)

        sheet8 = wb.create_sheet(title="文本语义解析")
        row_count = 0
        for sentenceindex in dict_sentenceindex_roleindex_argmeaningwordsdict:
            roleindex_argmeaningdict = dict_sentenceindex_roleindex_argmeaningwordsdict[sentenceindex]
            roleindex_argmeaningpostagsdict = dict_sentenceindex_roleindex_argmeaningpostagsdict[sentenceindex]
            roleindex_argmeaningarcsdict = dict_sentenceindex_roleindex_argmeaningarcsdict[sentenceindex]
            sentence = list_origion_text[sentenceindex]
            sentence_word_segmentation = list_word_segmentation[sentenceindex]
            sentence_arcs = list_arcs_content[sentenceindex]
            sentence_postags = list_postags[sentenceindex]
            row_count = row_count + 1
            sheet8.cell(row=row_count, column=1, value=sentence)

            column_count = 1
            for roleindex, argmeaningdict in roleindex_argmeaningdict.items():
                predicate = sentence_word_segmentation[roleindex]
                str_predicate_argmeaning_listscopewords = predicate + " : " + str(argmeaningdict)
                column_count = column_count + 1
                sheet8.cell(row=row_count, column=column_count, value=str_predicate_argmeaning_listscopewords)
            for roleindex, argmeaningarcsdict in roleindex_argmeaningarcsdict.items():
                predicate_arc = sentence_arcs[roleindex]
                str_predicate_argmeaning_listscopearcs = predicate_arc + " : " + str(argmeaningarcsdict)
                column_count = column_count + 1
                sheet8.cell(row=row_count, column=column_count, value=str_predicate_argmeaning_listscopearcs)
            for roleindex, argmeaningpostagsdict in roleindex_argmeaningpostagsdict.items():
                predicate_postag = sentence_postags[roleindex]
                str_predicate_argmeaning_listscopepostags = predicate_postag + " : " + str(argmeaningpostagsdict)
                column_count = column_count + 1
                sheet8.cell(row=row_count, column=column_count, value=str_predicate_argmeaning_listscopepostags)
        sheet9 = wb.create_sheet(title="文本语法解析")
        row_count = 0
        for sentenceindex in dict_sentenceindex_roleindex_argmeaningpostagsdict:
            sentence_word_segmentation = list_word_segmentation[sentenceindex]
            sentence_postags = list_postags[sentenceindex]
            sentence_arcs = list_arcs_content[sentenceindex]
            row_count = row_count + 1
            sheet9.cell(row_count, 1, value=str(sentence_word_segmentation))
            sheet9.cell(row_count, 2, value=str(sentence_arcs))
            sheet9.cell(row_count, 3, value=str(sentence_postags))

        row_count = 0
        sheet10 = wb.create_sheet(title="语义贡献度")
        for sentenceindex in range(len(list_origion_text)):
            sentence = list_origion_text[sentenceindex]
            contribute = list_contribute[sentenceindex]
            row_count = row_count + 1
            sheet10.cell(row_count, 1, value=str(sentence))
            sheet10.cell(row_count, 2, value=str(contribute))

        path = os.path.join(os.getcwd(), "excel")
        folder = os.path.exists(path)
        if not folder:
            os.makedirs(path)
        pathwrite = os.path.join(path, filename + ".xlsx")
        wb.save(pathwrite)
        print("写入数据成功！")
        pass

    def show_ratio(self, dict_amount_ratio, filename, figurename, xlabel, ylabel):
        plt.rcParams['font.sans-serif'] = ['SimHei']  # 指定默认字体 SimHei为黑体
        plt.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号
        plt.figure(figurename)
        list_labels = []
        list_amount = []
        list_ratio = []
        for i in range(len(dict_amount_ratio)):
            key = dict_amount_ratio[i][0]
            amount = dict_amount_ratio[i][1][0]
            ratio = dict_amount_ratio[i][1][1]
            list_labels.append(key)
            list_amount.append(amount)
            list_ratio.append(ratio)
        length = len(list_labels)
        plt.bar(range(length), list_ratio, tick_label=list_labels)
        plt.xlabel(xlabel)
        plt.ylabel(ylabel)
        plt.title(filename + "_" + figurename)
        path = os.path.join(os.getcwd(),"fig")
        folder = os.path.exists(path)
        if not folder:
            os.makedirs(path)
        else:
            pathwrite = os.path.join(path, filename + "_" + figurename + ".png")
            plt.savefig(pathwrite)
            plt.show()

    def function_normalized(self, x):
        y = 1 / (1 + math.exp(-0.3 * x))
        return y

    def ratio_normalized(self, ratio):
        return self.function_normalized(1/ratio)

    def semantic_contribution(self, datainfo, dict_sentenceindex_roleindex_argmeaningarcsdict,
                              dict_sentenceindex_roleindex_argmeaningpostagsdict):
        # 将arg_arcs_ratio和arg_postags_ratio转化为字典
        dict_arcs_singleratio = {}
        dict_postags_singleratio = {}
        dict_argmeaning_ratio = {}
        for tuple_arcs_ratio in datainfo.arg_arcs_ratio:
            key = tuple_arcs_ratio[0]
            value = tuple_arcs_ratio[1]
            dict_arcs_singleratio.update({key : value})
        for tuple_postags_ratio in datainfo.arg_postags_ratio:
            key = tuple_postags_ratio[0]
            value = tuple_postags_ratio[1]
            dict_postags_singleratio.update({key : value})
        for tuple_argmeaning_ratio in datainfo.arg_argmeaning_ratio:
            key = tuple_argmeaning_ratio[0]
            value = tuple_argmeaning_ratio[1]
            dict_argmeaning_ratio.update({key : value})

        dict_argmeaning_arcs_ratio = {}
        for index_argmeaningarcs in range(len(datainfo.arg_argmeaning_arcs_ratio)):
            argmeaning_arcs_ratio = datainfo.arg_argmeaning_arcs_ratio[index_argmeaningarcs]
            argmeaning = argmeaning_arcs_ratio[0]
            list_arcs_ratio = argmeaning_arcs_ratio[1]
            dict_arcs_ratio = {}
            for i in range(len(list_arcs_ratio)):
                arc = list_arcs_ratio[i][0]
                amount = list_arcs_ratio[i][1][0]
                ratio = list_arcs_ratio[i][1][1]
                dict_arcs_ratio.update({arc : [amount, ratio]})
            dict_argmeaning_arcs_ratio.update({argmeaning : dict_arcs_ratio})

        dict_argmeaning_postags_ratio = {}
        for index_argmeaningpostags in range(len(datainfo.arg_argmeaning_postags_ratio)):
            argmeaning_postags_ratio = datainfo.arg_argmeaning_postags_ratio[index_argmeaningpostags]
            argmeaning = argmeaning_postags_ratio[0]
            list_postags_ratio = argmeaning_postags_ratio[1]
            dict_postags_ratio = {}
            for i in range(len(list_postags_ratio)):
                postag = list_postags_ratio[i][0]
                amount = list_postags_ratio[i][1][0]
                ratio = list_postags_ratio[i][1][1]
                dict_postags_ratio.update({postag : [amount, ratio]})
            dict_argmeaning_postags_ratio.update({argmeaning : dict_postags_ratio})

        list_sentence_result = []
        for sentenceindex, dict_roleindex_argmeaningarcsdict in dict_sentenceindex_roleindex_argmeaningarcsdict.items():
            sentence_roleindex_argmeaningarcsdict = dict_sentenceindex_roleindex_argmeaningarcsdict[sentenceindex]
            sentence_roleindex_argmeaningpostagsdict = dict_sentenceindex_roleindex_argmeaningpostagsdict[sentenceindex]
            dict_argmeaning_result = {}
            for roleindex, dict_argmeaning_arcs in sentence_roleindex_argmeaningarcsdict.items():
                dict_argmeaning_arcs = sentence_roleindex_argmeaningarcsdict[roleindex]
                dict_argmeaning_postags = sentence_roleindex_argmeaningpostagsdict[roleindex]
                for argmeaning, list_arcs in dict_argmeaning_arcs.items():
                    list_arcs = dict_argmeaning_arcs[argmeaning]
                    list_postags = dict_argmeaning_postags[argmeaning]
                    dict_arcspostags_result = {} # 属于一个语义角色的计算结果
                    for index_arcspostags in range(len(list_arcs)):
                        arcs = list_arcs[index_arcspostags]
                        postags = list_postags[index_arcspostags]
                        for index_arcpostag in range(len(arcs)):
                            arc = arcs[index_arcpostag]
                            postag = postags[index_arcpostag]

                            # 原来计划使用的是这个，但是特征不强烈
                            # ratio_arc_origion = dict_arcs_singleratio[arc]
                            # ratio_postag_origion = dict_postags_singleratio[postag]

                            # 现在使用的是更为细分的dict
                            ratio_arc_origion = dict_argmeaning_arcs_ratio[argmeaning][arc]
                            ratio_postag_origion = dict_argmeaning_postags_ratio[argmeaning][postag]

                            normalized_arc_ratio = self.ratio_normalized(ratio_arc_origion[1])
                            normalized_postag_ratio = self.ratio_normalized(ratio_postag_origion[1])
                            normalized_arcmulpostag = normalized_arc_ratio * normalized_postag_ratio
                            if arc in dict_arcspostags_result:
                                temp = dict_arcspostags_result[arc]
                                temp[0] = temp[0] + normalized_arcmulpostag
                                temp[1] = temp[1] + 1
                                dict_arcspostags_result.update({arc : temp})
                            else:
                                temp = [normalized_arcmulpostag, 1]
                                dict_arcspostags_result.update({arc : temp})
                    arcspostags = 0
                    count_arcspostags = 0
                    for arc, list_arcratio in dict_arcspostags_result.items():
                        arcspostags = arcspostags + list_arcratio[0]/list_arcratio[1]
                        count_arcspostags = count_arcspostags + 1
                    arcspostags = arcspostags / count_arcspostags

                    argratio = dict_argmeaning_ratio[argmeaning]
                    argmularcspostags = arcspostags * self.ratio_normalized(argratio[1])
                    if argmeaning in dict_argmeaning_result:
                        temp = dict_argmeaning_result[argmeaning]
                        temp[0] = temp[0] + argmularcspostags
                        temp[1] = temp[1] + 1
                        dict_argmeaning_result.update({argmeaning : temp})
                    else:
                        temp = [argmularcspostags, 1]
                        dict_argmeaning_result.update({argmeaning : temp})
            count_sentence_result = 0
            sentence_result = 0
            for argmeaning, list_argmeaningresult in dict_argmeaning_result.items():
                sentence_result = sentence_result + list_argmeaningresult[0]/list_argmeaningresult[1]
                count_sentence_result = count_sentence_result + 1
            if count_sentence_result != 0:
                sentence_result = sentence_result / count_sentence_result
            list_sentence_result.append(sentence_result)
        return list_sentence_result

    '''
    {
        predicate: [[sentenceindex1, contribute], [sentenceindex2, contribute]...]
    }
    '''
    def downsize_by_predicate(self,
                              list_origion_corpus,
                              list_contribute,
                              dict_sentenceindex_roleindex_argmeaningwordsdict,
                              list_word_segmentation,
                              proportion):
        list_predicatesentence_result = []
        dict_predicate_listsentenceindex = {} # 存储每个谓词对应的[句子序号，句子贡献度]
        dict_predicate_listcontribute = {} # 存储每个谓词对应的[句子贡献度，出现次数]
        dict_predicate_contribute = {}
        dict_predicate_indexrecord = {}
        for sentenceindex, roleindex_argmeaningwordsdict in dict_sentenceindex_roleindex_argmeaningwordsdict.items():
            list_words = list_word_segmentation[sentenceindex]
            for roleindex in roleindex_argmeaningwordsdict.keys():
                predicate = list_words[roleindex]
                if predicate in dict_predicate_listsentenceindex:
                    list_temp = dict_predicate_listsentenceindex[predicate]
                    flag = 0
                    for i in range(len(list_temp)):
                        if sentenceindex == list_temp[i][0]:
                            flag = 1
                    if flag == 0:
                        list_temp.append([sentenceindex, list_contribute[sentenceindex]])
                        dict_predicate_listsentenceindex.update({predicate: list_temp})
                else:
                    dict_predicate_listsentenceindex.update({predicate: [[sentenceindex, list_contribute[sentenceindex]]]})

                if predicate in dict_predicate_listcontribute:
                    list_temp = dict_predicate_listcontribute[predicate]
                    list_record = dict_predicate_indexrecord[predicate]
                    flag = 0
                    for i in range(len(list_record)):
                        if sentenceindex == list_record[i]:
                            flag = 1
                    if flag == 0:
                        list_temp[0] = list_temp[0] + list_contribute[sentenceindex]
                        list_temp[1] = list_temp[1] + 1
                        dict_predicate_listcontribute.update({predicate: list_temp})
                        list_record.append(sentenceindex)
                        dict_predicate_indexrecord.update({predicate: list_record})
                else:
                    dict_predicate_listcontribute.update({predicate: [list_contribute[sentenceindex], 1]})
                    dict_predicate_indexrecord.update({predicate: [sentenceindex]})
        for predicate, list_contribute in dict_predicate_listcontribute.items():
            dict_predicate_contribute.update({predicate: list_contribute[0]/list_contribute[1]})
        list_predicate_contribute = sorted(dict_predicate_contribute.items(), key=lambda item: item[1], reverse=True)

        list_sentenceindex_record = []
        for tuple_predicate_contribute in list_predicate_contribute:
            predicate = tuple_predicate_contribute[0]
            list_sentenceindex = dict_predicate_listsentenceindex[predicate]
            list_sentenceindex = sorted(list_sentenceindex, key=lambda item: item[1], reverse=True)
            amount = int(math.ceil(proportion * len(list_sentenceindex)))
            for i in range(amount):
                sentenceindex = list_sentenceindex[i][0]
                flag = 0
                if sentenceindex in list_sentenceindex_record:
                    flag = 1
                if flag == 0:
                    list_predicatesentence_result.append(list_origion_corpus[sentenceindex])
                    list_sentenceindex_record.append(sentenceindex)
        return list_predicatesentence_result, list_predicate_contribute
        pass

    def downsize_by_args(self,
                         list_origion_corpus,
                         list_contribute,
                         dict_sentenceindex_roleindex_argmeaningwordsdict,
                         proportion):
        dict_argmeaning_sentenceindexlist = {}
        list_argmeaningsentence_result = []
        for sentenceindex, dict_roleindex_argmeaningwordsdict in dict_sentenceindex_roleindex_argmeaningwordsdict.items():
            for roleindex, argmeaningwordsdict in dict_roleindex_argmeaningwordsdict.items():
                for argmeaning, words in argmeaningwordsdict.items():
                    if argmeaning in dict_argmeaning_sentenceindexlist:
                        list_temp = dict_argmeaning_sentenceindexlist[argmeaning]
                        flag = 0
                        for i in range(len(list_temp)):
                            if sentenceindex == list_temp[i][0]:
                                flag = 1
                        if flag == 0:
                            list_temp.append([sentenceindex, list_contribute[sentenceindex]])
                            dict_argmeaning_sentenceindexlist.update({argmeaning: list_temp})
                    else:
                        dict_argmeaning_sentenceindexlist.update({argmeaning: [[sentenceindex, list_contribute[sentenceindex]]]})

        list_sentenceindex_record = []
        for argmeaning, list_sentenceindex in dict_argmeaning_sentenceindexlist.items():
            list_sentenceindex = sorted(list_sentenceindex, key=lambda item: item[1], reverse=True)
            amount = int(math.ceil(proportion * len(list_sentenceindex)))
            for i in range(amount):
                index = list_sentenceindex[i][0]
                flag = 0
                if index in list_sentenceindex_record:
                    flag = 1
                if flag == 0:
                    sentence = list_origion_corpus[index]
                    list_argmeaningsentence_result.append(sentence)
                    list_sentenceindex_record.append(index)
        return list_argmeaningsentence_result

    def downsize_by_contribute(self, list_origion_corpus, list_contribute, proportion):
        dict_sentence_contribute = {}
        list_sentence_downsize_by_contribute = []
        amount = int(math.ceil(proportion * len(list_origion_corpus)))
        for i in range(len(list_origion_corpus)):
            dict_sentence_contribute.update({list_origion_corpus[i]: list_contribute[i]})
        list_sentence_contribute = sorted(dict_sentence_contribute.items(), key=lambda item: item[1], reverse=True)
        for i in range(amount):
            list_sentence_downsize_by_contribute.append(list_sentence_contribute[i][0])
        return list_sentence_downsize_by_contribute
        pass

    def random_filter(self, dict_filename_corpus, ratio):
        dict_name_corpus_randomfilter = {}
        for filename, list_corpus in dict_filename_corpus.items():
            list_corpus_return = []
            for i in range(len(list_corpus)):
                sentence_strip = list_corpus[i].strip()
                if sentence_strip != "":
                    list_corpus_return.append(sentence_strip)
            random_num = int(len(list_corpus_return) * ratio)
            list_corpus_random = random.sample(list_corpus_return, random_num)
            dict_name_corpus_randomfilter.update({filename:list_corpus_random})
        return dict_name_corpus_randomfilter

    def writefilelist(self, dict_filename_list, path):
        for filename, filelist in dict_filename_list.items():
            filepath = os.path.join(path, filename)
            file_write = codecs.open(filepath, 'w', 'utf-8')
            for sentence in filelist:
                file_write.write(sentence + os.linesep)
            file_write.flush()
            file_write.close()

    def deduplicate_zh(self, dict_filename_filelist, coincidence_rate):
        dict_filename_filelist_deduplicate = {}
        for filename, filelist in dict_filename_filelist.items():
            filelist_deduplicate = []
            for i in range(len(filelist)):
                tag = 0
                for j in range(i+1, len(filelist)):
                    set_sentence_pre = set(list(filelist[i]))
                    set_sentence_back = set(list(filelist[j]))
                    set_intersection = set_sentence_pre & set_sentence_back
                    num_min = min(len(set_sentence_pre), len(set_sentence_back))
                    if len(set_intersection)/num_min >= coincidence_rate:
                        tag = 1
                if tag != 1:
                    filelist_deduplicate.append(filelist[i])
            dict_filename_filelist_deduplicate.update({filename:filelist_deduplicate})
        return dict_filename_filelist_deduplicate

if __name__ == '__main__':
    path_dir = "E:\\dailywork\\语法分析\\expriment4"
    SRL = SRL()
    dict_path_wordsegmentation = {}
    dict_path_postags = {}
    dict_path_arcs = {}
    dict_path_roles = {}

    dict_filename_corpus = SRL.readfilelist(path_dir)
    for filename in dict_filename_corpus.keys():
        list_corpus = dict_filename_corpus[filename]
        list_word_segmentation = SRL.word_segmentation(list_corpus)
        list_postags = SRL.postagger(list_word_segmentation)
        [list_arcs, list_arcs_content] = SRL.syntax_parser(list_word_segmentation, list_postags)
        list_roles = SRL.sementic_role_label(list_word_segmentation, list_postags, list_arcs)

        dict_path_wordsegmentation.update({filename:list_word_segmentation})
        dict_path_postags.update({filename:list_postags})
        dict_path_arcs.update({filename:list_arcs})
        dict_path_roles.update({filename:list_roles})
        dict_sentenceindex_roleindex_argdict = SRL.formatSRL(list_roles)

        [dict_sentenceindex_roleindex_argmeaningwordsdict,
         dict_sentenceindex_roleindex_argmeaningpostagsdict,
         dict_sentenceindex_roleindex_argmeaningarcsdict,
         list_core_words,
         list_core_words_postags,
         list_core_words_arcs,
         list_core_words_arcs_meaning] = SRL.parseSRL(dict_sentenceindex_roleindex_argdict,
                                                      list_word_segmentation,
                                                      list_postags,
                                                      list_arcs_content)
        [dict_arc_wordscount,
         dict_word_count,
         dict_postag_wordscount,
         dict_argmeaning_wordscount,
         dict_arc_amount,
         dict_postag_amount,
         dict_argmeaning_arcscountdict,
         dict_argmeaning_postagscountdict] = SRL.stasticalSRL(dict_sentenceindex_roleindex_argmeaningwordsdict,
                                                              dict_sentenceindex_roleindex_argmeaningpostagsdict,
                                                              dict_sentenceindex_roleindex_argmeaningarcsdict,
                                                              list_core_words,
                                                              list_core_words_postags,
                                                              list_core_words_arcs,
                                                              list_core_words_arcs_meaning)

        # 谓词统计量化指标
        # dict_arcmeaning_ratio = 谓词的语法角色数/谓词总数
        # dict_postags_ratio = 谓词的词性数/谓词总数
        # dict_word_ratio = 谓词数目/谓词总数
        [predicate_arcmeaning_ratio,
         predicate_postags_ratio,
         predicate_words_ratio] = SRL.ratio_predicate(dict_arc_wordscount,
                                                dict_postag_wordscount,
                                                dict_word_count)

        # 语义角色相关统计量化指标
        # dict_argmeaning_ratio = 语义角色数/词语总数
        # dict_arc_ratio = 语法角色数/词语总数
        # dict_postag_ratio = 词性数/词语总数
        # dict_argmeaning_arcsratiodict = 在语义角色范围内的语法角色数/语义角色范围内的词语总数
        # dict_argmeaning_postagsratiodict = 在语义角色范围内的词性数目/语义角色范围内的词语总数
        [arg_argmeaning_ratio,
         arg_arcs_ratio,
         arg_postags_ratio,
         arg_argmeaning_arcsratiodict,
         arg_argmeaning_postagsratiodict] = SRL.ratio_argument(dict_argmeaning_wordscount,
                                                                dict_arc_amount,
                                                                dict_postag_amount,
                                                                dict_argmeaning_arcscountdict,
                                                                dict_argmeaning_postagscountdict)

        datainfo = DataInfo(predicate_arcmeaning_ratio,
                            predicate_postags_ratio,
                            predicate_words_ratio,
                            arg_argmeaning_ratio,
                            arg_arcs_ratio,
                            arg_postags_ratio,
                            arg_argmeaning_arcsratiodict,
                            arg_argmeaning_postagsratiodict)

        list_contribute = SRL.semantic_contribution(datainfo, dict_sentenceindex_roleindex_argmeaningarcsdict,
                                                    dict_sentenceindex_roleindex_argmeaningpostagsdict)

        [list_predicatesentence_result,
         list_predicate_contribute] = SRL.downsize_by_predicate(list_corpus,
                                                                list_contribute,
                                                                dict_sentenceindex_roleindex_argmeaningwordsdict,
                                                                list_word_segmentation,
                                                                0.1)

        list_argmeaningsentence_result = SRL.downsize_by_args(list_corpus,
                                                              list_contribute,
                                                              dict_sentenceindex_roleindex_argmeaningwordsdict,
                                                              0.1)

        list_contribute_result = SRL.downsize_by_contribute(list_corpus, list_contribute, 0.1)

        dict_filename_random_result = SRL.random_filter({filename + "_randomfilter": list_corpus}, 0.1)

        dict_filename_predicatefilter = SRL.deduplicate_zh({filename + "_predicatefilter": list_predicatesentence_result}, 0.6)
        dict_filename_argmeaningfilter = SRL.deduplicate_zh({filename + "_argmeaningfilter": list_argmeaningsentence_result}, 0.6)
        dict_filename_contributefilter = SRL.deduplicate_zh({filename + "_contributefilter": list_contribute_result}, 0.6)
        dict_filename_randomfilter = SRL.deduplicate_zh(dict_filename_random_result, 0.6)

        path = os.path.join(os.getcwd(), "filter")
        folder = os.path.exists(path)
        if not folder:
            os.makedirs(path)
        SRL.writefilelist(dict_filename_predicatefilter, path)
        SRL.writefilelist(dict_filename_argmeaningfilter, path)
        SRL.writefilelist(dict_filename_contributefilter, path)
        SRL.writefilelist(dict_filename_randomfilter, path)

        SRL.writeXLS(datainfo,
                     list_corpus,
                     dict_sentenceindex_roleindex_argmeaningwordsdict,
                     dict_sentenceindex_roleindex_argmeaningpostagsdict,
                     dict_sentenceindex_roleindex_argmeaningarcsdict,
                     list_word_segmentation,
                     list_postags,
                     list_arcs_content,
                     list_contribute,
                     filename)

        SRL.show_ratio(datainfo.predicate_arcmeaning_ratio, filename, "谓词语法树角色比例", "arcs", "ratio")
        SRL.show_ratio(datainfo.predicate_postags_ratio, filename, "谓词词性比例", "postag", "ratio")
        SRL.show_ratio(datainfo.arg_argmeaning_ratio, filename, "语义角色比例", "args", "ratio")
        SRL.show_ratio(datainfo.arg_arcs_ratio, filename, "语义角色语法树比例", "arg-scope-total arcs", "ratio")
        SRL.show_ratio(datainfo.arg_postags_ratio, filename, "语义角色词性比例", "arg-scope postag", "ratio")